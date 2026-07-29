"""
CDP scraper — connects to existing Edge browser, scrapes Alibaba product pages.
Uses Playwright locators (not regex) for reliable element selection.
"""
import time, re, json, os, sys, random
import openpyxl
from playwright.sync_api import sync_playwright

EXCEL_PATH = r"C:\Users\ZhuanZ（无密码）\Downloads\Beauty_and_Personal_Care_Trend_High_Profit_Products_2026.xlsx"
CDP_URL = "http://127.0.0.1:9222"
OUTPUT_DIR = r"E:\meng-feifei\project-alone-website\awl-vertex-glow-marker\src\content\products"


def scrape_tiered_pricing(page):
    """Extract tiered/volume pricing from Alibaba detail page.

    Returns list of {'qty': int, 'price': float} dicts, or empty list.
    Prices are converted from CNY to USD if needed.
    """
    tiers = []

    try:
        # Alibaba 2026 ladder price: [data-testid="ladder-price"] > .price-item
        ladder = page.locator('[data-testid="ladder-price"]')
        if ladder.count() == 0:
            # Fallback: .module_price .price-item
            ladder = page.locator('.module_price .price-item').first
            if ladder.count() > 0:
                ladder = page.locator('.module_price')

        if ladder.count() > 0:
            items = ladder.first.locator('.price-item')
            if items.count() == 0:
                # If no .price-item children, try direct text parsing
                raw_text = ladder.first.inner_text()
                return _parse_tier_text(raw_text)

            for i in range(items.count()):
                item = items.nth(i)
                text = item.inner_text()
                tier = _parse_single_tier(text)
                if tier:
                    tiers.append(tier)

            if tiers:
                # Convert CNY to USD if prices seem to be in CNY
                avg = sum(t['price'] for t in tiers) / len(tiers)
                if avg > 100:  # Likely in CNY (e.g., ¥150 vs $1.50), not JPY
                    pass  # Keep original for now, user browses from China
                return tiers
    except:
        pass

    # Fallback: parse from body text
    body_text = page.locator('body').inner_text()
    return _parse_tier_text(body_text)


def _parse_tier_text(text):
    """Parse tiered pricing from plain text. Handles both $ and ¥."""
    tiers = []
    # Match: "$X.XX  10-99 pieces" or "¥X.XX  500-1,999 pieces" or "≥100 pieces  $X.XX"
    currency = '$'
    if '¥' in text:
        currency = '¥'
    elif '￥' in text:
        currency = '￥'

    # Pattern: currency amount then quantity range
    lines = text.split('\n')
    for line in lines:
        line = line.strip()
        if not line or len(line) > 120:
            continue

        # Try: "¥2.70" + "500-1,999 pieces"
        price_m = re.search(r'[¥￥\$]\s*([\d,.]+)', line)
        qty_m = re.search(
            r'(?:≥|>=)?\s*(\d[\d,]*)\s*(?:[-–]\s*(\d[\d,]*))?\s*(?:pieces|pcs|units|piece|pc|unit)',
            line, re.IGNORECASE
        )

        if price_m and qty_m:
            price = float(price_m.group(1).replace(',', ''))
            qty_min = int(qty_m.group(1).replace(',', ''))
            if not any(t['qty'] == qty_min for t in tiers):
                tiers.append({'qty': qty_min, 'price': price})
            continue

        # Try: quantity then price "500-1,999 pieces ¥2.70"
        qty_first = re.match(
            r'(?:≥|>=)?\s*(\d[\d,]*)\s*(?:[-–]\s*\d[\d,]*)?\s*(?:pieces|pcs|units).*?[¥￥\$]\s*([\d,.]+)',
            line, re.IGNORECASE
        )
        if qty_first:
            qty_min = int(qty_first.group(1).replace(',', ''))
            price = float(qty_first.group(2).replace(',', ''))
            if not any(t['qty'] == qty_min for t in tiers):
                tiers.append({'qty': qty_min, 'price': price})

    return tiers


def _parse_single_tier(text):
    """Parse a single tier from .price-item text like '¥2.70\\n500-1,999 pieces'."""
    # Extract price
    price_m = re.search(r'[¥￥\$]\s*([\d,.]+)', text)
    if not price_m:
        return None

    price = float(price_m.group(1).replace(',', ''))

    # Extract quantity: "500-1,999 pieces", "≥5,000 pieces", "2,000-4,999 pieces"
    qty_m = re.search(
        r'(?:≥|>=)?\s*(\d[\d,]*)\s*(?:[-–]\s*\d[\d,]*)?\s*(?:pieces|pcs|units|piece|pc|unit)',
        text, re.IGNORECASE
    )
    if not qty_m:
        return None

    qty_min = int(qty_m.group(1).replace(',', ''))
    return {'qty': qty_min, 'price': price}

CAT_MAP = [
    (2, 7,  "nails", "Nail Supplies"),
    (8, 13, "eyelashes", "Eyelash Supplies"),
    (14, 19, "permanent-makeup", "Permanent Makeup"),
    (20, 25, "beauty-tools", "Beauty Tools"),
    (26, 31, "daily-chemical", "Daily Chemical"),
    (32, 37, "makeup-tools", "Makeup Tools"),
    (38, 43, "hair-tools", "Hair Tools"),
]

def get_category(row_idx):
    for start, end, slug, name in CAT_MAP:
        if start <= row_idx <= end:
            return slug, name
    return "beauty-tools", "Beauty Tools"

def generate_sku(cat_slug, idx):
    prefix_map = {
        "nails": "NAIL", "eyelashes": "LASH", "permanent-makeup": "PMUS",
        "beauty-tools": "BTL", "daily-chemical": "DCML", "makeup-tools": "MKTL",
        "hair-tools": "HAIR",
    }
    return f"{prefix_map.get(cat_slug, 'PROD')}-{idx:04d}"

def slugify(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text[:80].strip('-')

def read_excel():
    """Read product data from Excel, returns list of dicts."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[1]]
    products = []
    for row in range(2, ws.max_row + 1):
        idx = ws.cell(row, 1).value
        name_en = ws.cell(row, 4).value
        purchase_price = ws.cell(row, 7).value
        retail_price = ws.cell(row, 8).value
        margin = ws.cell(row, 9).value
        search_url = ws.cell(row, 10).value
        if not name_en:
            continue
        cat_slug, cat_name = get_category(row)
        products.append({
            'row': row, 'idx': idx,
            'name_en': str(name_en).strip(),
            'purchase_price': float(purchase_price) if purchase_price else 9.99,
            'retail_price': float(retail_price) if retail_price else None,
            'margin': str(margin).strip() if margin else '',
            'search_url': str(search_url).strip() if search_url else '',
            'cat_slug': cat_slug,
            'cat_name': cat_name,
        })
    wb.close()
    return products

def has_real_images(filepath):
    """Check if product MD has real scraped images (not pexels placeholders)."""
    if not os.path.exists(filepath):
        return False
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    return 'alicdn.com' in content


def scrape_product_from_search(page, product_info):
    """Navigate to search page, find best-selling product, go to detail, scrape."""
    search_url = product_info['search_url']
    if not search_url:
        print(f"  SKIP: no search URL")
        return None

    print(f"  Loading search: {search_url[:100]}...")
    try:
        page.goto(search_url, timeout=45000, wait_until='load')
    except Exception as e:
        # Timeout or redirect — page may still have loaded enough
        print(f"  Nav note: {str(e)[:80]}")
    time.sleep(4)

    # Check for captcha — auto-wait for user to solve in browser
    try:
        title = page.title()
    except:
        print(f"  Page context lost during CAPTCHA check, continuing...")
        title = ''
    if '验证' in title or 'Verification' in title or 'Security' in title:
        print(f"  CAPTCHA DETECTED! Waiting for it to be solved (checking every 3s, max 120s)...")
        for _ in range(40):
            time.sleep(3)
            try:
                new_title = page.title()
                if not ('验证' in new_title or 'Verification' in new_title or 'Security' in new_title):
                    print(f"  CAPTCHA solved! Continuing...")
                    time.sleep(2)
                    break
            except:
                pass
        else:
            print(f"  CAPTCHA not solved in time, skipping this product")
            return None

    # Find offerlist
    ol = page.locator('#sse-fluent-offerlist')
    try:
        ol.wait_for(state='visible', timeout=5000)
    except:
        print(f"  No offerlist found on this page")
        page.screenshot(path=f'no_offerlist_{product_info["idx"]}.png')
        return None

    # Find product-detail links using Playwright locators
    links = ol.locator('a[href*="product-detail"]')
    count = links.count()
    print(f"  Found {count} product links")

    if count == 0:
        # Try broader search
        links = page.locator('a[href*="product-detail"]')
        count = links.count()
        print(f"  Broader search: {count} links")

    if count == 0:
        print(f"  No product links found")
        return None

    # Find best-seller: look for sold counts
    best_idx = 0
    best_sold = 0

    # Get all link hrefs and nearby text
    for i in range(min(count, 20)):
        try:
            link = links.nth(i)
            # Get parent card text to find sold count
            parent_text = link.locator('xpath=ancestor::*[contains(@class, "offer") or contains(@class, "card") or contains(@class, "list")][1]').inner_text()
            sold_match = re.findall(r'(\d[\d,]*)\s*(?:sold|pieces|orders|transactions|solds)', parent_text, re.IGNORECASE)
            if sold_match:
                sold_num = int(sold_match[0].replace(',', ''))
                if sold_num > best_sold:
                    best_sold = sold_num
                    best_idx = i
        except:
            pass

    if best_sold == 0:
        print(f"  No sales data found, using first result")
        best_idx = 0

    # Click the best product
    try:
        best_link = links.nth(best_idx)
        href = best_link.get_attribute('href')
        # Clean URL
        product_url = re.sub(r'\?.*$', '', href)
        if not product_url.startswith('http'):
            product_url = 'https:' + product_url if product_url.startswith('//') else 'https://www.alibaba.com' + product_url
        print(f"  Selected (#{best_idx}, sold={best_sold}): {product_url[:120]}")
    except Exception as e:
        print(f"  Error getting link: {e}")
        return None

    # Navigate to detail page
    print(f"  Loading detail page...")
    try:
        page.goto(product_url, timeout=45000, wait_until='load')
    except Exception as e:
        print(f"  Detail nav note: {str(e)[:80]}")
    time.sleep(4)

    # Check captcha again
    try:
        title = page.title()
    except:
        title = ''
    if '验证' in title or 'Verification' in title:
        print(f"  CAPTCHA on detail page! Waiting (max 120s)...")
        for _ in range(40):
            time.sleep(3)
            try:
                new_title = page.title()
                if not ('验证' in new_title or 'Verification' in new_title):
                    print(f"  CAPTCHA solved!")
                    time.sleep(2)
                    break
            except:
                pass
        else:
            print(f"  CAPTCHA timeout, skipping")
            return None

    return scrape_detail(page, product_info)

def scrape_detail(page, product_info):
    """Scrape the current detail page."""
    result = {
        'title': '', 'price': 0, 'price_range': '',
        'sample_moq': '', 'custom_moq': '',
        'images': [], 'specs': [], 'pricing': [],
    }

    # Title
    try:
        h1 = page.locator('h1').first
        result['title'] = h1.inner_text().strip()
    except:
        try:
            result['title'] = page.locator('[data-module="product-title"]').first.inner_text().strip()
        except:
            result['title'] = product_info['name_en']

    # Images from product image thumbs
    try:
        thumbs = page.locator('[data-submodule="ProductImageThumbsList"]')
        if thumbs.count() > 0:
            slides = thumbs.locator('[aria-roledescription="slide"]')
            slide_count = slides.count()
            print(f"  Slides: {slide_count}")
            for i in range(min(slide_count, 10)):
                try:
                    slide = slides.nth(i)
                    inner = slide.locator('div[style*="background-image"]')
                    if inner.count() > 0:
                        style = inner.get_attribute('style') or ''
                        m = re.search(r'url\(["\']?(//[^"\')\s]+)["\']?\)', style)
                        if m:
                            url = 'https:' + m.group(1)
                            url = re.sub(r'_\d+x\d+\.', r'_500x500.', url)
                            if url not in result['images']:
                                result['images'].append(url)
                except:
                    pass
    except Exception as e:
        print(f"  Thumbs error: {e}")

    # Fallback: find all product images
    if not result['images']:
        imgs = page.locator('img[src*="alicdn.com"]')
        seen = set()
        for i in range(min(imgs.count(), 30)):
            try:
                src = imgs.nth(i).get_attribute('src') or ''
                if '/kf/' in src or 'product' in src.lower():
                    src = re.sub(r'_\d+x\d+\.', r'_500x500.', src)
                    if src not in seen:
                        seen.add(src)
                        result['images'].append(src)
            except:
                pass

    # Body text for regex scraping
    body_text = page.locator('body').inner_text()

    # Price
    pm = re.search(r'US\s*\$([\d,.]+)\s*-\s*US\s*\$([\d,.]+)', body_text)
    if not pm:
        pm = re.search(r'\$([\d,.]+)\s*-\s*\$([\d,.]+)', body_text)
    if pm:
        result['price_range'] = f'${pm.group(1)} - ${pm.group(2)}'
        result['price'] = float(pm.group(1).replace(',', ''))
    else:
        for pat in [r'US\s*\$([\d,.]+)', r'FOB\s*Price\s*:?\s*US\s*\$([\d,.]+)', r'\$([\d,.]+)']:
            pm = re.search(pat, body_text)
            if pm:
                result['price'] = float(pm.group(1).replace(',', ''))
                break

    if not result['price']:
        result['price'] = product_info['purchase_price']

    # Tiered pricing
    result['pricing'] = scrape_tiered_pricing(page)

    # MOQ
    moq_m = re.search(r'Min(?:imum)?\.?\s*[Oo]rder\s*:?\s*(\d[\d,]*\s*(?:Piece|Pieces|Set|Sets|Unit|Units|[Pp]cs|[Pp]c|Kilogram|kilograms|kg|Gram|grams|g|Ton|tons|Meter|meters|m|Liter|liters|L|Pair|pairs|Carton|cartons|Box|boxes|Bag|bags|Roll|rolls|Sheet|sheets|Bottle|bottles|Dozen|dozens)[\s\w]*)', body_text)
    if moq_m:
        result['sample_moq'] = moq_m.group(1).strip()

    # Custom/OEM MOQ
    custom_m = re.search(r'(?:Custom(?:ized|isation)?|Private\s*Label|OEM|ODM)\s*(?:MOQ|Min(?:imum)?\s*(?:Order|Qty)|Qty)\s*:?\s*(\d[\d,]*\s*(?:Piece|Pieces|Set|Sets|Unit|Units|[Pp]cs|[Pp]c|Kilogram|kilograms|kg)[\s\w]*)', body_text, re.IGNORECASE)
    if custom_m:
        result['custom_moq'] = custom_m.group(1).strip()

    # Attributes
    attr_keys = [
        'Brand Name', 'Model Number', 'Material', 'Color', 'Size', 'Type',
        'Feature', 'Usage', 'Power', 'Voltage', 'Weight', 'Certification',
        'Place of Origin', 'Supply Ability', 'Packaging', 'Port', 'Warranty',
        'After-sales Service', 'Product name', 'Function', 'Application',
        'Battery', 'Charging Time', 'Capacity', 'Dimensions', 'OEM/ODM',
        'Style', 'Gender', 'Age Group',
    ]
    for key in attr_keys:
        m = re.search(rf'{key}\s*:?\s*([^\n]+)', body_text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()[:200]
            if val and val.lower() not in ('n/a', 'none', '-', '', '--'):
                result['specs'].append({'label': key, 'value': val})

    # Table attributes
    try:
        tables = page.locator('table')
        for ti in range(min(tables.count(), 3)):
            rows = tables.nth(ti).locator('tr')
            for ri in range(rows.count()):
                cells = rows.nth(ri).locator('td, th')
                if cells.count() >= 2:
                    label = cells.nth(0).inner_text().strip()
                    value = cells.nth(1).inner_text().strip()
                    if label and value and len(label) < 80 and len(value) < 300:
                        if not any(s['label'] == label for s in result['specs']):
                            result['specs'].append({'label': label, 'value': value})
    except:
        pass

    print(f"  Title: {result['title'][:80]}")
    print(f"  Price: {result['price']}")
    print(f"  Tiered pricing: {len(result['pricing'])} tiers")
    print(f"  MOQ: {result['sample_moq']}")
    print(f"  Images: {len(result['images'])}")
    print(f"  Specs: {len(result['specs'])}")

    return result

def generate_md(product_info, scraped):
    """Generate MD file content from product info and scraped data."""
    cat_slug = product_info['cat_slug']
    cat_name = product_info['cat_name']
    name_en = product_info['name_en']
    idx = product_info['idx']

    sku = generate_sku(cat_slug, idx)
    slug = slugify(name_en)

    # Use scraped price or Excel price
    price = scraped.get('price', 0) or product_info['purchase_price']
    retail_price = product_info.get('retail_price')
    raw_tiers = scraped.get('pricing', [])

    # Detect CNY and convert to USD (CNY prices are typically > 10x USD prices)
    # Beauty wholesale items: nail drill ~$28 USD vs ¥203 CNY; gel polish ~$8.50 USD vs ¥61 CNY
    purchase_usd = product_info.get('purchase_price', 0)
    cny_to_usd = None
    if raw_tiers and purchase_usd > 0:
        min_tier = min(t['price'] for t in raw_tiers) if raw_tiers else 0
        if min_tier > purchase_usd * 5:  # Scraped price is way higher than USD purchase price
            cny_to_usd = purchase_usd / min_tier  # Calculate conversion factor
        elif min_tier < purchase_usd * 0.15:  # Min tier suspiciously low vs purchase price
            pass  # Keep as-is

    if cny_to_usd is None and price > 50:  # Prices > $50 uncommon for most beauty products
        cny_to_usd = 0.14  # Rough CNY→USD

    pricing_tiers = []
    for t in raw_tiers:
        usd_price = round(t['price'] * cny_to_usd, 2) if cny_to_usd else t['price']
        pricing_tiers.append({'qty': t['qty'], 'price': usd_price})

    if cny_to_usd and price == scraped.get('price'):
        price = round(price * cny_to_usd, 2)

    # Images
    images = scraped.get('images', [])
    main_img = images[0] if images else ''
    gallery_imgs = images[1:5] if len(images) > 1 else []

    # MOQ
    moq = scraped.get('sample_moq', '') or '10 pcs'

    # Badge from margin
    margin_val = float(str(product_info['margin']).replace('%', '')) if product_info['margin'] else 0
    if margin_val >= 95:
        badge = "Hot Seller"
    elif margin_val >= 90:
        badge = "Bestseller"
    elif margin_val >= 85:
        badge = "Trending"
    elif margin_val >= 80:
        badge = "Value Deal"
    else:
        badge = ""

    # Description
    desc_parts = [f"Professional {cat_name.lower()} product"]
    if retail_price:
        desc_parts.append(f"Retail: ${retail_price:.2f}")
    if product_info['margin']:
        desc_parts.append(f"Margin: {product_info['margin']}")
    desc_parts.append("Wholesale from our Guangzhou Baiyun factory.")
    description = " — ".join(desc_parts)

    # Use English Excel name as title, store scraped Chinese title as alt
    display_title = name_en
    scraped_title = scraped.get('title', '')

    # Build frontmatter
    md = f'''---
title: "{display_title}"
sku: "{sku}"
subtitle: "{cat_name} wholesale — high profit margin product"
category: "{cat_slug}"
subcategory: ""
price: {price:.2f}
stock: 9999
moq: "{moq}"
'''

    # Tiered pricing
    if pricing_tiers:
        md += "pricing:\n"
        for t in pricing_tiers:
            md += f'  - qty: {t["qty"]}\n    price: {t["price"]:.2f}\n'
    else:
        # Fallback: generate 3 tiers from the single price
        md += "pricing:\n"
        md += f'  - qty: 10\n    price: {price:.2f}\n'
        md += f'  - qty: 100\n    price: {price * 0.85:.2f}\n'
        md += f'  - qty: 1000\n    price: {price * 0.70:.2f}\n'

    md += f'image: "{main_img}"\n'

    if gallery_imgs:
        md += "gallery:\n"
        for g in gallery_imgs:
            md += f'  - image: "{g}"\n'

    if badge:
        md += f'badge: "{badge}"\n'

    # Specs
    md += "specs:\n"
    md += f'  - label: "Category"\n    value: "{cat_name}"\n'
    md += f'  - label: "Wholesale Price (USD)"\n    value: "${price:.2f}"\n'
    if retail_price:
        md += f'  - label: "Suggested Retail (USD)"\n    value: "${retail_price:.2f}"\n'
    if product_info['margin']:
        md += f'  - label: "Gross Margin"\n    value: "{product_info["margin"]}"\n'
    md += f'  - label: "MOQ"\n    value: "{moq}"\n'
    md += f'  - label: "Supply Ability"\n    value: "10000 pcs per week"\n'

    # Scraped specs
    for spec in scraped.get('specs', [])[:15]:
        label = spec['label'].replace('"', "'")
        value = spec['value'].replace('"', "'")
        md += f'  - label: "{label}"\n    value: "{value}"\n'

    md += "---\n\n"
    md += description + "\n"

    return slug, md

def main():
    print("=== Alibaba CDP Scraper (Playwright) ===")
    print(f"CDP: {CDP_URL}")
    print()

    # Read Excel
    products = read_excel()
    print(f"Loaded {len(products)} products from Excel\n")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(CDP_URL)
        context = browser.contexts[0]
        page = context.pages[0] if context.pages else context.new_page()

        print(f"Connected. Current URL: {page.url[:100]}")
        print()

        results = []

        for i, prod in enumerate(products):
            cat_slug = prod['cat_slug']
            name_en = prod['name_en']
            slug = slugify(name_en)
            filepath = os.path.join(OUTPUT_DIR, f"{slug}.md")
            if has_real_images(filepath):
                print(f"[{i+1}/{len(products)}] {cat_slug}/{name_en[:60]} — SKIP (has real images)")
                results.append({'product': name_en, 'status': 'skip_exists'})
                continue

            print(f"[{i+1}/{len(products)}] {cat_slug}/{name_en[:60]}")

            try:
                scraped = scrape_product_from_search(page, prod)
                if scraped:
                    slug, md_content = generate_md(prod, scraped)
                    filepath = os.path.join(OUTPUT_DIR, f"{slug}.md")
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(md_content)
                    print(f"  -> Saved: {filepath}")
                    results.append({'product': prod['name_en'], 'status': 'ok', 'images': len(scraped['images'])})
                else:
                    print(f"  -> FAILED (no data)")
                    results.append({'product': prod['name_en'], 'status': 'no_data'})
            except Exception as e:
                print(f"  -> ERROR: {e}")
                results.append({'product': prod['name_en'], 'status': f'error: {e}'})

            print()
            # Random delay to avoid triggering captcha
            delay = random.uniform(4, 8)
            time.sleep(delay)

        # Summary
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        ok = sum(1 for r in results if r['status'] == 'ok')
        print(f"Success: {ok}/{len(results)}")
        for r in results:
            print(f"  [{r['status']}] {r['product'][:70]}")

        # Save detailed results
        with open('scrape_results.json', 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        print("\nSaved scrape_results.json")

if __name__ == '__main__':
    main()
