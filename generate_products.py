"""
Generate 42 product MD files from Excel data.
Uses English names, prices, margins. Pexels for placeholder images.
"""
import openpyxl
import re
import os

EXCEL_PATH = r"C:\Users\ZhuanZ（无密码）\Downloads\Beauty_and_Personal_Care_Trend_High_Profit_Products_2026.xlsx"
OUTPUT_DIR = r"E:\meng-feifei\project-alone-website\awl-vertex-glow-marker\src\content\products"

# Row ranges -> category
CAT_MAP = [
    (2, 7,  "nails", "Nail Supplies"),
    (8, 13, "eyelashes", "Eyelash Supplies"),
    (14, 19, "permanent-makeup", "Permanent Makeup"),
    (20, 25, "beauty-tools", "Beauty Tools"),
    (26, 31, "daily-chemical", "Daily Chemical"),
    (32, 37, "makeup-tools", "Makeup Tools"),
    (38, 43, "hair-tools", "Hair Tools"),
]

# Product images by category (Pexels)
CAT_IMAGES = {
    "nails": [
        "https://images.pexels.com/photos/3997388/pexels-photo-3997388.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3428102/pexels-photo-3428102.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3326936/pexels-photo-3326936.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3326929/pexels-photo-3326929.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/15756184/pexels-photo-15756184.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3997394/pexels-photo-3997394.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "eyelashes": [
        "https://images.pexels.com/photos/3993446/pexels-photo-3993446.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3373745/pexels-photo-3373745.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4586686/pexels-photo-4586686.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4587665/pexels-photo-4587665.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993448/pexels-photo-3993448.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3738338/pexels-photo-3738338.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "permanent-makeup": [
        "https://images.pexels.com/photos/8864652/pexels-photo-8864652.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993449/pexels-photo-3993449.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4587666/pexels-photo-4587666.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993310/pexels-photo-3993310.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4612146/pexels-photo-4612146.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/7755550/pexels-photo-7755550.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "beauty-tools": [
        "https://images.pexels.com/photos/5069610/pexels-photo-5069610.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/5938624/pexels-photo-5938624.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3738347/pexels-photo-3738347.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/5938557/pexels-photo-5938557.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4587669/pexels-photo-4587669.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993309/pexels-photo-3993309.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "daily-chemical": [
        "https://images.pexels.com/photos/4465124/pexels-photo-4465124.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4041392/pexels-photo-4041392.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/4202924/pexels-photo-4202924.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3738344/pexels-photo-3738344.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/5938363/pexels-photo-5938363.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/5938558/pexels-photo-5938558.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "makeup-tools": [
        "https://images.pexels.com/photos/2533266/pexels-photo-2533266.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3373743/pexels-photo-3373743.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/2115373/pexels-photo-2115373.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/1403633/pexels-photo-1403633.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3373737/pexels-photo-3373737.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/1082519/pexels-photo-1082519.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
    "hair-tools": [
        "https://images.pexels.com/photos/8467997/pexels-photo-8467997.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3738345/pexels-photo-3738345.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/973402/pexels-photo-973402.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993311/pexels-photo-3993311.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993323/pexels-photo-3993323.jpeg?auto=compress&cs=tinysrgb&w=600",
        "https://images.pexels.com/photos/3993298/pexels-photo-3993298.jpeg?auto=compress&cs=tinysrgb&w=600",
    ],
}

def slugify(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text[:80].strip('-')


def get_category(row_idx):
    for start, end, slug, name in CAT_MAP:
        if start <= row_idx <= end:
            return slug, name
    return "beauty-tools", "Beauty Tools"


def generate_sku(cat_slug, idx):
    prefix_map = {
        "nails": "NAIL", "eyelashes": "LASH", "permanent-makeup": "PMUS",
        "beauty-tools": "BTL", "daily-chemical": "DCML", "makeup-tools": "MKTL",
        "hair-tools": "HAIR",
    }
    prefix = prefix_map.get(cat_slug, "PROD")
    return f"{prefix}-{idx:04d}"


print("Reading Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[1]]

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Track existing slugs to avoid duplicates
existing_slugs = set()
for f in os.listdir(OUTPUT_DIR):
    if f.endswith('.md'):
        existing_slugs.add(f.replace('.md', ''))
print(f"Found {len(existing_slugs)} existing product files, will only generate missing ones.")

count = 0
for row in range(2, ws.max_row + 1):
    idx = ws.cell(row, 1).value
    name_en = ws.cell(row, 4).value
    purchase_price = ws.cell(row, 7).value
    retail_price = ws.cell(row, 8).value
    margin = ws.cell(row, 9).value

    if not name_en:
        continue

    cat_slug, cat_name = get_category(row)

    # Images
    img_list = CAT_IMAGES.get(cat_slug, CAT_IMAGES["beauty-tools"])
    main_img = img_list[(row - 2) % len(img_list)]
    gallery_imgs = [img_list[(row - 2 + i) % len(img_list)] for i in range(1, 4)]

    # Slug, SKU
    slug = slugify(name_en)
    sku = generate_sku(cat_slug, idx)

    # Skip if already exists
    if slug in existing_slugs:
        count += 1
        print(f"[{count:2d}] SKIP (exists) {cat_slug:20s} | {name_en[:55]}")
        continue

    # Price
    price = float(purchase_price) if purchase_price else 9.99
    stock = 9999
    moq = "10 pcs"

    # Badge based on margin
    margin_val = float(str(margin).replace('%', '')) if margin else 0
    if margin_val >= 95:
        badge = "Hot Seller"
    elif margin_val >= 90:
        badge = "Bestseller"
    elif margin_val >= 85:
        badge = "Trending"
    elif margin_val >= 80:
        badge = "Value Deal"
    else:
        badge = ""

    # Build description
    desc_parts = [f"Professional {cat_name.lower()} product"]
    if retail_price:
        desc_parts.append(f"Retail: ${float(retail_price):.2f}")
    if margin:
        desc_parts.append(f"Margin: {str(margin).strip()}")
    desc_parts.append("Wholesale from our Guangzhou Baiyun factory.")
    description = " — ".join(desc_parts)

    # Build MD
    md = f'''---
title: "{name_en}"
sku: "{sku}"
subtitle: "{cat_name} wholesale — high profit margin product"
category: "{cat_slug}"
subcategory: ""
price: {price:.2f}
stock: {stock}
moq: "{moq}"
image: "{main_img}"
gallery:
  - image: "{gallery_imgs[0]}"
  - image: "{gallery_imgs[1]}"
  - image: "{gallery_imgs[2]}"
'''

    if badge:
        md += f'badge: "{badge}"\n'

    # Specs
    md += "specs:\n"
    md += f'  - label: "Category"\n    value: "{cat_name}"\n'
    md += f'  - label: "Wholesale Price (USD)"\n    value: "${price:.2f}"\n'
    if retail_price:
        md += f'  - label: "Suggested Retail (USD)"\n    value: "${float(retail_price):.2f}"\n'
    if margin:
        md += f'  - label: "Gross Margin"\n    value: "{str(margin).strip()}"\n'
    md += f'  - label: "MOQ"\n    value: "{moq}"\n'
    md += f'  - label: "Supply Ability"\n    value: "10000 pcs per week"\n'

    md += "---\n\n"
    md += description + "\n"

    filepath = os.path.join(OUTPUT_DIR, f"{slug}.md")
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(md)

    count += 1
    print(f"[{count:2d}] {cat_slug:20s} | {sku:12s} | ${price:7.2f} | {name_en[:55]}")

print(f"\nDone! {count} products total ({count - len(existing_slugs)} newly created).")
